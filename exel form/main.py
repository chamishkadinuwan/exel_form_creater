import openpyxl
from openpyxl import load_workbook
from tkinter import Tk, Label, Entry, Button


# Function to update the Excel sheet with provided data
def update_excel(name, age, telephone):
    workbook_path = 'data.xlsx'  # Change this to your desired Excel file path
    sheet_name = 'Sheet'

    # Load the workbook or create a new one if it doesn't exist
    try:
        workbook = load_workbook(workbook_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Access the sheet or create a new one if it doesn't exist
    sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)

    # Append data to the sheet
    sheet.append([name, age, telephone])

    # Save the workbook
    workbook.save(workbook_path)


# Function to handle button click and update Excel sheet
def submit_form():
    name = name_entry.get()
    age = age_entry.get()
    telephone = telephone_entry.get()

    # Update the Excel sheet with the provided data
    update_excel(name, age, telephone)

    # Clear the entry fields after submission
    name_entry.delete(0, 'end')
    age_entry.delete(0, 'end')
    telephone_entry.delete(0, 'end')


# Create the GUI
root = Tk()
root.title('Excel Table Updater')

# Labels and Entry fields for name, age, and telephone number
Label(root, text='Name:').grid(row=0, column=0)
name_entry = Entry(root)
name_entry.grid(row=0, column=1)

Label(root, text='Age:').grid(row=1, column=0)
age_entry = Entry(root)
age_entry.grid(row=1, column=1)

Label(root, text='Telephone:').grid(row=2, column=0)
telephone_entry = Entry(root)
telephone_entry.grid(row=2, column=1)

# Submit button to update the Excel sheet
submit_button = Button(root, text='Submit', command=submit_form)
submit_button.grid(row=3, column=0, columnspan=2)

# Run the GUI
root.mainloop()
